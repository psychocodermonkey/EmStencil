"""
 Program: Centralized .xlsx read/write for template import and export.
    Name: Andrew Dixon            File: spreadsheet.py
    Date: 3 Apr 2026
   Notes:

   Copyright (c) 2023-2026 Andrew Dixon

  This file is part of EmStencil.
  Licensed under the GNU Lesser General Public License v2.1.
  See the LICENSE file at the project root for details.
........1.........2.........3.........4.........5.........6.........7.........8.........9.........0.........1.........2.........3..
"""

from __future__ import annotations

from zipfile import BadZipFile
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from .content_html import exportContentAsHTML
from .Exceptions import InvalidImportFileType


EXPORT_HEADERS: tuple[str, str, str] = ('Title', 'Content', 'Tags')


def _cellStr(value: object) -> str:
  if value is None:
    return ''

  return str(value).strip()


def readTemplateRows(path: str) -> list[tuple[str, str, list[str]]]:
  """
  Load the first worksheet in the workbook. Row 1 is skipped (header).
  Columns A–C are title, content, and comma-separated tags (split only; normalize elsewhere).
  """
  try:
    wb = load_workbook(path, read_only=True, data_only=True)

  except (BadZipFile, InvalidFileException, OSError) as e:
    raise InvalidImportFileType() from e

  try:
    if not wb.worksheets:
      raise InvalidImportFileType()

    ws = wb.worksheets[0]
    out: list[tuple[str, str, list[str]]] = []

    for row in ws.iter_rows(min_row=2, max_col=3, values_only=True):
      title = _cellStr(row[0] if row else None)
      content = _cellStr(row[1] if row and len(row) > 1 else None)
      rawTags = row[2] if row and len(row) > 2 else None
      tagsCell = _cellStr(rawTags) if rawTags is not None else ''
      tagParts = tagsCell.split(',') if tagsCell else []
      out.append((title, content, tagParts))

  finally:
    wb.close()

  return out


def writeTemplatesWorkbook(path: str, rows: list[tuple[str, str, str]]) -> None:
  """Write a new workbook; Content column is always HTML (plain bodies wrapped on export)."""
  wb = Workbook()
  ws = wb.active
  ws.append(list(EXPORT_HEADERS))

  for title, content, tagsCsv in rows:
    ws.append([title, exportContentAsHTML(content), tagsCsv])

  wb.save(path)
