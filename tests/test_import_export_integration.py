#! /usr/bin/env python3

"""
 Program: Import/export integration — DB ↔ xlsx round-trips and convertSpreadsheet.
    Name: Andrew Dixon            File: test_import_export_integration.py
    Date: 5 Apr 2026
   Notes:

  Copyright (c) 2023-2026 Andrew Dixon

  This file is part of EmStencil.
  Licensed under the GNU Lesser General Public License v2.1.
  See the LICENSE file at the project root for details.
........1.........2.........3.........4.........5.........6.........7.........8.........9.........0.........1.........2.........3..
"""

from __future__ import annotations

from pathlib import Path

import pytest
from openpyxl import Workbook

from emstencil.content_html import is_html_content
from emstencil.Database import TemplateDB
from emstencil.Dataclasses import EmailTemplate, MetadataTag
from emstencil.ImportTemplates import convertSpreadsheet
from emstencil.spreadsheet import EXPORT_HEADERS, read_template_rows, write_templates_workbook


def _content_for_title(templateDB: TemplateDB, title: str) -> str:
  cursor = templateDB.getConnection().cursor()
  cursor.execute('select content from templates where title = ?;', [title])
  row = cursor.fetchone()
  assert row is not None
  return row[0]


def testExportThenImportUpsertsPlainTemplateAsHtmlInDb(templateDB: TemplateDB, tmp_path: Path) -> None:
  """App export wraps plain Content in HTML; re-import stores that string in SQLite."""
  t = EmailTemplate('Zebra', 'Hello ${x} & y')
  t.metadata = [MetadataTag('one')]
  templateDB.AddTemplate(t)
  path = tmp_path / 'round.xlsx'
  write_templates_workbook(str(path), templateDB.FetchAllTemplatesForExport())
  convertSpreadsheet(str(path), templateDB)
  assert _content_for_title(templateDB, 'Zebra') == '<p>Hello ${x} &amp; y</p>'
  reparsed = EmailTemplate('Zebra', _content_for_title(templateDB, 'Zebra'))
  assert list(reparsed.fields) == ['x']
  assert is_html_content(reparsed.content) is True


def testExportThenImportPreservesHtmlTemplateBody(templateDB: TemplateDB, tmp_path: Path) -> None:
  body = '<table><tr><td>${cell}</td></tr></table>'
  t = EmailTemplate('HtmlT', body)
  t.metadata = [MetadataTag('t')]
  templateDB.AddTemplate(t)
  path = tmp_path / 'h.xlsx'
  write_templates_workbook(str(path), templateDB.FetchAllTemplatesForExport())
  convertSpreadsheet(str(path), templateDB)
  assert _content_for_title(templateDB, 'HtmlT') == body
  round_trip = EmailTemplate('x', _content_for_title(templateDB, 'HtmlT'))
  assert list(round_trip.fields) == ['cell']


def testImportPlainXlsxColumnBStoredWithoutExportWrapper(templateDB: TemplateDB, tmp_path: Path) -> None:
  """Files not produced by write_templates_workbook can still ship plain Content."""
  path = tmp_path / 'legacy.xlsx'
  wb = Workbook()
  ws = wb.active
  ws.append(list(EXPORT_HEADERS))
  ws.append(['Legacy', 'Line1\nLine2 ${id}', 'alpha,beta'])
  wb.save(path)
  convertSpreadsheet(str(path), templateDB)
  assert _content_for_title(templateDB, 'Legacy') == 'Line1\nLine2 ${id}'
  assert is_html_content(_content_for_title(templateDB, 'Legacy')) is False
  et = EmailTemplate('Legacy', _content_for_title(templateDB, 'Legacy'))
  assert et.fields == {'id': None}


def testConvertSpreadsheetDuplicateTitlesInWorkbookRaises(templateDB: TemplateDB, tmp_path: Path) -> None:
  path = tmp_path / 'dup.xlsx'
  wb = Workbook()
  ws = wb.active
  ws.append(list(EXPORT_HEADERS))
  ws.append(['Same', '<p>a</p>', ''])
  ws.append(['Same', '<p>b</p>', ''])
  wb.save(path)
  with pytest.raises(ValueError, match='Duplicate template titles'):
    convertSpreadsheet(str(path), templateDB)


def testReadTemplateRowsMatchesWriteTemplatesWorkbook(tmp_path: Path) -> None:
  """Guards column layout contract between export and import."""
  path = tmp_path / 'grid.xlsx'
  rows_in = [('T1', 'plain ${a}', 'x'), ('T2', '<p>${b}</p>', 'y,z')]
  write_templates_workbook(str(path), rows_in)
  assert read_template_rows(str(path)) == [
    ('T1', '<p>plain ${a}</p>', ['x']),
    ('T2', '<p>${b}</p>', ['y', 'z']),
  ]
